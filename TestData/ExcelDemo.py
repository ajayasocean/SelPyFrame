import openpyxl

test_data = {}
book = openpyxl.load_workbook("/Users/ajaysagar/ocean/SelPyFrame/TestData/pyxlDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
# print(cell.value)
# print(sheet['A5'].value)
print("Rows: ", sheet.max_row)
# print(sheet.max_column)

for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == 'TC1':
        for j in range(2, sheet.max_column):
            # print(sheet.cell(row=i, column=j).value)
            test_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(test_data)

