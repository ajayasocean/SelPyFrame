import openpyxl


class HomeData:

    test_home_data = [{'name': 'Tester', 'email': 'qa@tester.com', 'password': 'test@123', 'gender': 'Male', 'dob': '02/02/1993'},
                        {'name': 'qa', 'email': 'test@qa.com', 'password': 'test@1234', 'gender': 'Female', 'dob': '03/02/1993'}
                        ]

    @staticmethod
    def get_test_data(test_case_name):
        test_data = {}
        book = openpyxl.load_workbook("/Users/ajaysagar/ocean/SelPyFrame/TestData/pyxlDemo.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column):
                    # print(sheet.cell(row=i, column=j).value)
                    test_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [test_data]

